import streamlit as st
import pandas as pd
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from datetime import date
import streamlit.components.v1 as components
import base64
import os
BRAND_YELLOW = "#FFD700"
GRAPHITE = "#111111"
LIGHT_GRAY = "#F5F5F5"

# =========================
# LOGO / BACKGROUND
# =========================
def get_base64(path):
    try:
        with open(path, "rb") as f:
            return base64.b64encode(f.read()).decode()
    except:
        return ""

bg_logo = get_base64("assets/fs_logo_white.png")

# =========================
# BRANDING
# =========================
def apply_branding(page_title, subtitle):
    st.markdown(f"""
    <style>
    html, body, [class*="css"] {{
        font-family: 'Nunito Sans', 'Segoe UI', sans-serif;
    }}

    .stApp {{
        background-color: {LIGHT_GRAY};
        overflow-x: hidden;
    }}

    .stApp::before {{
        content: "";
        position: fixed;
        inset: -300px;
        background-image: url("data:image/png;base64,{bg_logo}");
        background-size: 340px;
        background-repeat: repeat;
        background-position: 0 0;
        opacity: 0.045;
        transform: rotate(-18deg);
        z-index: 0;
        pointer-events: none;
        animation: bgMove 55s linear infinite;
    }}

    @keyframes bgMove {{
        from {{ background-position: 0 0; }}
        to {{ background-position: 900px 900px; }}
    }}

    .block-container {{
        position: relative;
        z-index: 1;
    }}

    .brand-header {{
        background: rgba(17,17,17,0.96);
        padding: 22px 28px;
        border-radius: 18px;
        margin-bottom: 26px;
        border-left: 10px solid {BRAND_YELLOW};
        box-shadow: 0 8px 24px rgba(0,0,0,0.18);
    }}

    .brand-title {{
        color: white;
        font-size: 32px;
        font-weight: 900;
        margin: 0;
    }}

    .brand-subtitle {{
        color: #d9d9d9;
        font-size: 15px;
        margin-top: 4px;
    }}

    .stTextInput input,
    .stNumberInput input,
    .stDateInput input {{
        background-color: white !important;
        color: black !important;
        border: 1px solid #d0d0d0 !important;
        border-radius: 10px !important;
    }}

    div[data-baseweb="select"] > div {{
        background-color: white !important;
        color: black !important;
        border-radius: 10px !important;
        border: 1px solid #d0d0d0 !important;
    }}

    div[data-baseweb="select"] svg {{
        display: none !important;
    }}

    div.stButton > button,
    button[kind="secondary"],
    [data-testid="stNumberInput"] button,
    [data-testid="stNumberInput"] div button {{
        background: {GRAPHITE} !important;
        color: white !important;
        border: 1px solid {GRAPHITE} !important;
        border-radius: 10px !important;
        font-weight: 700 !important;
    }}

    div.stButton > button:hover,
    button[kind="secondary"]:hover,
    [data-testid="stNumberInput"] button:hover,
    [data-testid="stNumberInput"] div button:hover {{
        background: black !important;
        color: {BRAND_YELLOW} !important;
        border: 1px solid {BRAND_YELLOW} !important;
    }}

    .stDownloadButton > button {{
        background: {BRAND_YELLOW} !important;
        color: black !important;
        border-radius: 10px !important;
        font-weight: 800 !important;
        border: none !important;
    }}

    [data-testid="stDataFrame"] {{
        background-color: white !important;
        border-radius: 12px;
        padding: 6px;
    }}
    </style>

    <div class="brand-header">
        <div class="brand-title">{page_title}</div>
        <div class="brand-subtitle">{subtitle}</div>
    </div>
    """, unsafe_allow_html=True)

# =========================
# HEADER
# =========================
col_logo, col_title = st.columns([1, 5])

with col_logo:
    try:
        st.image("assets/fs_logo.png", width=120)
    except:
        pass

with col_title:
    apply_branding(
        "CMDB Pregled",
        "Pretraga opreme i interni prenos između sektora"
    )

# =========================
# SESSION STATE
# =========================
if "transfer_list" not in st.session_state:
    st.session_state.transfer_list = []

if "generated_excel" not in st.session_state:
    st.session_state.generated_excel = None

if "generated_file_name" not in st.session_state:
    st.session_state.generated_file_name = ""

if "print_html" not in st.session_state:
    st.session_state.print_html = ""

if "last_chance_results" not in st.session_state:
    st.session_state.last_chance_results = pd.DataFrame()

if "search_triggered" not in st.session_state:
    st.session_state.search_triggered = False

if "main_table_key" not in st.session_state:
    st.session_state.main_table_key = 0

if "advanced_table_key" not in st.session_state:
    st.session_state.advanced_table_key = 0

# =========================
# LOAD DATA
# =========================
@st.cache_data
def load_data():
    try:
        return pd.read_excel("data.xlsx", dtype=str).fillna("")
    except:
        return pd.DataFrame()

df = load_data()

if df.empty:
    st.warning("data.xlsx nije pronađen ili je prazan")
    st.stop()

# =========================
# HELPERS
# =========================
def find_config_item_column(dataframe):
    possible_names = [
        "Config Item",
        "ConfigItem",
        "Config item",
        "Configuration Item",
        "CI",
        "Number"
    ]

    for col in possible_names:
        if col in dataframe.columns:
            return col

    return None


CONFIG_ITEM_COL = find_config_item_column(df)


def get_display_columns(dataframe):
    base_cols = [
        "Name",
        "Config Item",
        "Vendor",
        "Model",
        "Type",
        "SPInventoryNumber",
        "InventoryNumber",
        "SerialNumber"
    ]

    if CONFIG_ITEM_COL and CONFIG_ITEM_COL != "Config Item":
        base_cols = [
            "Name",
            CONFIG_ITEM_COL,
            "Vendor",
            "Model",
            "Type",
            "SPInventoryNumber",
            "InventoryNumber",
            "SerialNumber"
        ]

    return [c for c in base_cols if c in dataframe.columns]



def set_cell(ws, cell, value):
    for merged_range in ws.merged_cells.ranges:
        if cell in merged_range:
            top_left = merged_range.start_cell.coordinate
            ws[top_left] = value
            ws[top_left].alignment = Alignment(horizontal="center", vertical="center")
            return

    ws[cell] = value
    ws[cell].alignment = Alignment(horizontal="center", vertical="center")


def to_excel(dataframe):
    output = BytesIO()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        dataframe.to_excel(writer, index=False, sheet_name="CMDB")

    return output.getvalue()


def get_logo_base64():
    logo_path = "assets/fs_logo_print.png"

    if not os.path.exists(logo_path):
        logo_path = "assets/fs_logo.png"

    try:
        with open(logo_path, "rb") as f:
            return base64.b64encode(f.read()).decode()
    except:
        return ""


def build_print_html(selected_rows, transfer_type):
    if transfer_type == "BG_NS":
        broj_prenosa = "BG-NS"
        iz_magacina = "FSBG"
        zaduzio = "FSNS"
    else:
        broj_prenosa = "FSNIS-FSNS"
        iz_magacina = "FSNIŠ"
        zaduzio = "FSNS"

    logo_b64 = get_logo_base64()
    logo_html = ""

    if logo_b64:
        logo_html = f'<img src="data:image/png;base64,{logo_b64}" class="logo">'

    rows_html = ""

    for i in range(25):
        if i < len(selected_rows):
            row = selected_rows[i]

            rows_html += f"""
            <tr>
                <td>{i+1}</td>
                <td>{row.get("Name", "")}</td>
                <td>{row.get("Model", "")}</td>
                <td>{row.get("InventoryNumber", "")}</td>
                <td>{row.get("SerialNumber", "")}</td>
                <td>{row.get("SPInventoryNumber", "")}</td>
                <td></td>
            </tr>
            """
        else:
            rows_html += f"""
            <tr>
                <td>{i+1}</td>
                <td></td><td></td><td></td><td></td><td></td><td></td>
            </tr>
            """

    return f"""
    <html>
    <head>
        <style>
            @page {{
                size: A4 landscape;
                margin: 10mm;
            }}

            body {{
                font-family: Arial, sans-serif;
                color: #000;
                font-size: 12px;
                margin: 0;
                padding: 0;
            }}

            .print-btn {{
                margin-bottom: 12px;
                padding: 8px 16px;
                background: #2C2C2C;
                color: white;
                border: none;
                border-radius: 6px;
                font-weight: bold;
                cursor: pointer;
            }}

            table {{
                border-collapse: collapse;
                width: 100%;
            }}

            td, th {{
                border: 1px solid #000;
                padding: 4px;
                text-align: center;
                vertical-align: middle;
                height: 22px;
            }}

            .top-grid {{
                display: grid;
                grid-template-columns: 180px 1fr 320px;
                align-items: start;
                margin-bottom: 10px;
            }}

            .logo {{
                max-width: 130px;
                max-height: 70px;
            }}

            .company {{
                font-size: 13px;
                padding-top: 8px;
            }}

            .doc-title {{
                font-weight: bold;
                font-size: 15px;
                text-align: right;
                padding-top: 30px;
            }}

            .doc-box,
            .date-box {{
                display: inline-block;
                min-width: 110px;
                border-bottom: 1px solid #000;
                text-align: center;
                margin-left: 8px;
            }}

            .meta {{
                margin-top: 8px;
                width: 60%;
            }}

            .meta td {{
                height: 22px;
                padding: 3px;
            }}

            .meta-label {{
                font-weight: bold;
                text-align: left;
                background: #f5f5f5;
            }}

            .big-value {{
                font-size: 18px;
                font-weight: bold;
            }}

            .items {{
                margin-top: 14px;
            }}

            .items th {{
                font-weight: bold;
                background: #f5f5f5;
                height: 24px;
            }}

            .signatures {{
                display: grid;
                grid-template-columns: 1fr 1fr;
                gap: 120px;
                margin-top: 50px;
                padding: 0 80px;
            }}

            .sig {{
                text-align: center;
                border-top: 1px solid #000;
                padding-top: 6px;
                font-weight: bold;
            }}

            @media print {{
                .print-btn {{
                    display: none;
                }}
            }}
        </style>
    </head>

    <body>
        <button class="print-btn" onclick="window.print()">🖨️ Print</button>

        <div class="top-grid">
            <div>{logo_html}</div>
            <div class="company">Fiscal Solutions DOO, Temerinska 102, 21000 Novi Sad</div>
            <div class="doc-title">
                OTPREMNICA BR. <span class="doc-box">{broj_prenosa}</span><br>
                <span style="font-weight:normal;">Datum:</span>
                <span class="date-box">{date.today().strftime("%d.%m.%Y")}</span>
            </div>
        </div>

        <table class="meta">
            <tr>
                <td class="meta-label">Iz magacina / Ime i prezime</td>
                <td class="meta-label">UREĐAJ ZADUŽIO (ime i prezime / naziv firme)</td>
            </tr>
            <tr>
                <td class="big-value">{iz_magacina}</td>
                <td class="big-value">{zaduzio}</td>
            </tr>
        </table>

        <table class="items">
            <tr>
                <th>BR</th>
                <th>NAZIV</th>
                <th>MODEL</th>
                <th>INV</th>
                <th>SN</th>
                <th>SP/FS</th>
                <th>NAPOMENA</th>
            </tr>
            {rows_html}
        </table>

        <div class="signatures">
            <div class="sig">Robu izdao</div>
            <div class="sig">Robu primio</div>
        </div>
    </body>
    </html>
    """


def generate_internal_transfer(selected_rows, transfer_type):
    if not selected_rows:
        st.error("Lista je prazna.")
        return

    if transfer_type == "BG_NS":
        broj_prenosa = "BG-NS"
        iz_magacina = "FSBG"
        uredjaj_zaduzio = "FSNS"
        file_name = "interni_prenos_BG_NS.xlsx"
        iz_magacina_cell = "B8"
    else:
        broj_prenosa = "FSNIS-FSNS"
        iz_magacina = "FSNIŠ"
        uredjaj_zaduzio = "FSNS"
        file_name = "interni_prenos_NIS_NS.xlsx"
        iz_magacina_cell = "C8"

    try:
        wb = load_workbook("otpremnica_template.xlsx")
        ws = wb.active
    except:
        st.error("Nije pronađen fajl: otpremnica_template.xlsx")
        return

    set_cell(ws, "F4", broj_prenosa)
    set_cell(ws, "G5", date.today().strftime("%d.%m.%Y"))
    set_cell(ws, iz_magacina_cell, iz_magacina)
    set_cell(ws, "G8", uredjaj_zaduzio)

    for i, row in enumerate(selected_rows, start=1):
        r = 14 + i - 1

        set_cell(ws, f"B{r}", i)
        set_cell(ws, f"C{r}", row.get("Name", ""))
        set_cell(ws, f"D{r}", row.get("Model", ""))
        set_cell(ws, f"E{r}", row.get("InventoryNumber", ""))
        set_cell(ws, f"F{r}", row.get("SerialNumber", ""))
        set_cell(ws, f"G{r}", row.get("SPInventoryNumber", ""))

    output = BytesIO()
    wb.save(output)

    st.session_state.generated_excel = output.getvalue()
    st.session_state.generated_file_name = file_name
    st.session_state.print_html = build_print_html(selected_rows, transfer_type)


def add_selected(selected_rows):
    if selected_rows.empty:
        st.warning("Nisi izabrao nijedan uređaj.")
        return

    added = 0

    existing_sp = [
        x.get("SPInventoryNumber", "")
        for x in st.session_state.transfer_list
    ]

    for _, row in selected_rows.iterrows():
        sp = row.get("SPInventoryNumber", "")

        if sp and sp not in existing_sp:
            st.session_state.transfer_list.append(row.to_dict())
            existing_sp.append(sp)
            added += 1

    if added > 0:
        st.success(f"Dodato uređaja: {added}")
    else:
        st.warning("Nema novih uređaja za dodavanje.")


def build_search_results():
    filters = {
        "SPInventoryNumber": st.session_state.get("search_sp", "").strip(),
        "InventoryNumber": st.session_state.get("search_inv", "").strip(),
        "SerialNumber": st.session_state.get("search_serial", "").strip(),
        "Name": st.session_state.get("search_name", "").strip(),
        "Vendor": st.session_state.get("search_vendor", "").strip(),
        "Model": st.session_state.get("search_model", "").strip(),
        "Type": st.session_state.get("search_type", "").strip(),
    }

    active_filters = {
        col: val
        for col, val in filters.items()
        if val and col in df.columns
    }

    if not active_filters:
        return pd.DataFrame()

    result = df.copy()

    for col, val in active_filters.items():
        result = result[
            result[col]
            .astype(str)
            .str.contains(val, case=False, na=False)
        ]

    return result


# =========================
# SEARCH
# =========================
st.markdown("---")
st.subheader("🔎 Pretraga")

with st.form("search_form"):
    c1, c2, c3, c4, c5, c6, c7 = st.columns(7)

    with c1:
        st.text_input("SPInventoryNumber", key="search_sp")

    with c2:
        st.text_input("InventoryNumber", key="search_inv")

    with c3:
        st.text_input("SerialNumber", key="search_serial")

    with c4:
        st.text_input("Name", key="search_name")

    with c5:
        st.text_input("Vendor", key="search_vendor")

    with c6:
        st.text_input("Model", key="search_model")

    with c7:
        st.text_input("Type", key="search_type")

    search_clicked = st.form_submit_button("🔎 Pretraži")

if search_clicked:
    st.session_state.search_triggered = True
    st.session_state.main_table_key += 1

if st.session_state.search_triggered:
    filtered_df = build_search_results()

    st.markdown("---")
    st.subheader(f"📦 Rezultati: {len(filtered_df)}")

    if filtered_df.empty:
        st.info("Nema rezultata.")
    else:
        available_cols = get_display_columns(filtered_df)
        view_df = filtered_df[available_cols].copy()

        event = st.dataframe(
            view_df,
            use_container_width=True,
            hide_index=True,
            height=300,
            key=f"main_results_table_{st.session_state.main_table_key}",
            on_select="rerun",
            selection_mode="multi-row"
        )

        selected_indices = event.selection.rows

        selected_indices = [
            i for i in selected_indices
            if 0 <= i < len(view_df)
        ]

        selected = (
            view_df.iloc[selected_indices]
            if len(selected_indices) > 0
            else pd.DataFrame()
        )

        if st.button("➕ Dodaj uređaj"):
            add_selected(selected)

        st.download_button(
            "📥 Preuzmi filtrirani CMDB",
            data=to_excel(view_df),
            file_name="cmdb_pregled.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

# =========================
# NAPREDNA PRETRAGA
# =========================
st.markdown("---")

left_lc, middle_lc, right_lc = st.columns([1, 2, 1])

with middle_lc:
    st.subheader("🔍 Napredna pretraga")
    st.caption("Pretraga po svim dostupnim kolonama CMDB baze")

    last = st.text_input(
        "Pretraga po svim kolonama",
        key="last_chance_search"
    )

    search_last_clicked = st.button("🔎 Pokreni naprednu pretragu")

if search_last_clicked:
    st.session_state.advanced_table_key += 1
    if last:
        excluded = [
            "Description",
            "Owner",
            "WarrantyExpirationDate",
            "WarrantzExpirationDate",
            "InstallDate",
            "Note"
        ]

        search_df = df.drop(
            columns=[c for c in excluded if c in df.columns],
            errors="ignore"
        )

        mask = search_df.apply(
            lambda row: row.astype(str).str.lower().str.contains(last.lower()).any(),
            axis=1
        )

        st.session_state.last_chance_results = df[mask].copy()
    else:
        st.session_state.last_chance_results = pd.DataFrame()

if not st.session_state.last_chance_results.empty:
    st.markdown("---")
    st.subheader(
        f"📦 Rezultati napredne pretrage: {len(st.session_state.last_chance_results)}"
    )

    available_cols = get_display_columns(st.session_state.last_chance_results)

    view_df = st.session_state.last_chance_results[available_cols].copy()

    event_last = st.dataframe(
        view_df,
        use_container_width=True,
        hide_index=True,
        height=300,
        key=f"last_chance_results_table_{st.session_state.advanced_table_key}",
        on_select="rerun",
        selection_mode="multi-row"
    )

    selected_last_indices = event_last.selection.rows

    selected_last_indices = [
        i for i in selected_last_indices
        if 0 <= i < len(view_df)
    ]

    selected_last = (
        view_df.iloc[selected_last_indices]
        if len(selected_last_indices) > 0
        else pd.DataFrame()
    )

    if st.button("➕ Dodaj iz napredne pretrage"):
        add_selected(selected_last)

# =========================
# TRANSFER LIST
# =========================
st.markdown("---")
st.subheader("🔁 Lista za interni prenos")

if st.session_state.transfer_list:
    header = st.columns([2, 2, 2, 2, 2, 1])

    header[0].markdown("**Name**")
    header[1].markdown("**Model**")
    header[2].markdown("**SP**")
    header[3].markdown("**Inventory**")
    header[4].markdown("**Serial**")
    header[5].markdown("**Ukloni**")

    for i, row in enumerate(st.session_state.transfer_list):
        c = st.columns([2, 2, 2, 2, 2, 1])

        c[0].write(row.get("Name", ""))
        c[1].write(row.get("Model", ""))
        c[2].write(row.get("SPInventoryNumber", ""))
        c[3].write(row.get("InventoryNumber", ""))
        c[4].write(row.get("SerialNumber", ""))

        if c[5].button("🗑️", key=f"del{i}"):
            st.session_state.transfer_list.pop(i)
            st.rerun()

    st.info(f"Ukupno uređaja za prenos: {len(st.session_state.transfer_list)}")
else:
    st.info("Lista je prazna.")

# =========================
# ACTIONS
# =========================
col1, col2, col3 = st.columns(3)

with col1:
    if st.button("BG → NS"):
        generate_internal_transfer(
            st.session_state.transfer_list,
            "BG_NS"
        )

with col2:
    if st.button("NIŠ → NS"):
        generate_internal_transfer(
            st.session_state.transfer_list,
            "NIS_NS"
        )

with col3:
    if st.button("Obriši listu"):
        st.session_state.transfer_list = []
        st.session_state.generated_excel = None
        st.session_state.generated_file_name = ""
        st.session_state.print_html = ""
        st.rerun()

# =========================
# DOWNLOAD + PRINT
# =========================
if st.session_state.generated_excel:
    st.markdown("---")
    st.subheader("📄 Dokument")

    d1, d2 = st.columns(2)

    with d1:
        st.download_button(
            "📥 Download Excel",
            st.session_state.generated_excel,
            st.session_state.generated_file_name,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    with d2:
        if st.button("🖨️ Print dokument"):
            components.html(
                st.session_state.print_html,
                height=900,
                scrolling=True
            )
