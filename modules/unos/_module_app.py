import streamlit as st
import pandas as pd
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from datetime import date
import base64
BRAND_YELLOW = "#FFD700"
GRAPHITE = "#111111"
LIGHT_GRAY = "#F5F5F5"

# =========================
# BARCODE HELPERS
# =========================
def apply_barcode_pending(key: str):
    pending_key = f"{key}__barcode_pending"
    if pending_key in st.session_state:
        st.session_state[key] = st.session_state[pending_key]
        del st.session_state[pending_key]


def decode_barcode_from_camera(image_file):
    if image_file is None:
        return ""

    try:
        from PIL import Image
        import numpy as np
        from pyzbar.pyzbar import decode
    except Exception:
        return "__MISSING_BARCODE_LIBS__"

    try:
        image = Image.open(image_file)
        decoded = decode(np.array(image))
        if decoded:
            return decoded[0].data.decode("utf-8", errors="ignore").strip()
    except Exception:
        return ""

    return ""


def barcode_camera(label: str, key: str):
    with st.expander(f"📷 Skeniraj barkod za {label}", expanded=False):
        image = st.camera_input(
            "Usmeri kameru ka barkodu i slikaj",
            key=f"{key}__camera",
        )

        if image is not None:
            scanned = decode_barcode_from_camera(image)

            if scanned == "__MISSING_BARCODE_LIBS__":
                st.error("Nedostaju barcode biblioteke. U requirements.txt dodaj: Pillow, numpy, pyzbar. Na Linux serveru treba i libzbar0.")
            elif scanned:
                st.success(f"Pročitano: {scanned}")
                st.session_state[f"{key}__barcode_pending"] = scanned
                st.rerun()
            else:
                st.warning("Barkod nije pročitan. Probaj bliže, pod boljim svetlom ili pod drugim uglom.")


def barcode_text_input(label: str, key: str, **kwargs):
    apply_barcode_pending(key)
    value = st.text_input(label, key=key, **kwargs)
    barcode_camera(label, key)
    return value


def get_base64(path):
    try:
        with open(path, "rb") as f:
            return base64.b64encode(f.read()).decode()
    except:
        return ""

bg_logo = get_base64("assets/fs_logo_white.png")

def apply_branding(page_title, subtitle):
    st.markdown(f"""
    <style>
    html, body, [class*="css"] {{
        font-family: 'Nunito Sans', 'Segoe UI', sans-serif;
    }}

    .stApp {{
        background-color: {LIGHT_GRAY};
        overflow-x: hidden;
    }}

    .stApp::before {{
        content: "";
        position: fixed;
        inset: -300px;
        background-image: url("data:image/png;base64,{bg_logo}");
        background-size: 340px;
        background-repeat: repeat;
        background-position: 0 0;
        opacity: 0.045;
        transform: rotate(-18deg);
        z-index: 0;
        pointer-events: none;
        animation: bgMove 55s linear infinite;
    }}

    @keyframes bgMove {{
        from {{ background-position: 0 0; }}
        to {{ background-position: 900px 900px; }}
    }}

    .block-container {{
        position: relative;
        z-index: 1;
        max-width: 980px !important;
        padding-left: 2rem !important;
        padding-right: 2rem !important;
    }}

    .brand-header {{
        background: rgba(17,17,17,0.96);
        padding: 22px 28px;
        border-radius: 18px;
        margin-bottom: 26px;
        border-left: 10px solid {BRAND_YELLOW};
        box-shadow: 0 8px 24px rgba(0,0,0,0.18);
    }}

    .brand-title {{
        color: white;
        font-size: 32px;
        font-weight: 900;
        margin: 0;
    }}

    .brand-subtitle {{
        color: #d9d9d9;
        font-size: 15px;
        margin-top: 4px;
    }}

    .stTextInput input,
    .stNumberInput input,
    .stDateInput input {{
        background-color: white !important;
        color: black !important;
        border: 1px solid #d0d0d0 !important;
        border-radius: 10px !important;
    }}

    div[data-baseweb="select"] > div {{
        background-color: white !important;
        color: black !important;
        border-radius: 10px !important;
        border: 1px solid #d0d0d0 !important;
    }}

    div[data-baseweb="select"] svg {{
        display: none !important;
    }}

    div[data-baseweb="select"] [aria-label="open"] {{
        display: none !important;
    }}

    div.stButton > button,
    button[kind="secondary"],
    [data-testid="stNumberInput"] button,
    [data-testid="stNumberInput"] div button {{
        background: {GRAPHITE} !important;
        color: white !important;
        border: 1px solid {GRAPHITE} !important;
        border-radius: 10px !important;
        font-weight: 700 !important;
    }}

    div.stButton > button:hover,
    button[kind="secondary"]:hover,
    [data-testid="stNumberInput"] button:hover,
    [data-testid="stNumberInput"] div button:hover {{
        background: black !important;
        color: {BRAND_YELLOW} !important;
        border: 1px solid {BRAND_YELLOW} !important;
    }}

    .stDownloadButton > button {{
        background: {BRAND_YELLOW} !important;
        color: black !important;
        border-radius: 10px !important;
        font-weight: 800 !important;
        border: none !important;
    }}
    </style>

    <div class="brand-header">
        <div class="brand-title">{page_title}</div>
        <div class="brand-subtitle">{subtitle}</div>
    </div>
    """, unsafe_allow_html=True)

col_logo, col_title = st.columns([1, 5])

with col_logo:
    try:
        st.image("assets/fs_logo.png", width=120)
    except:
        pass

with col_title:
    apply_branding("CMDB Unos", "Unos uređaja i generisanje prijemnice / otpremnice")

if "doc_type" not in st.session_state:
    st.session_state.doc_type = None

DEFAULT_TYPE_OPTIONS = [
    "Desktop", "Laptop", "Cash drawer", "Cradle",
    "IP Phone", "Monitor", "Monitor Touch Screen",
    "Printer Pos", "Printer label", "Router",
    "Switch", "Scanner Counter", "Scanner Hand",
    "Scanner Terminal", "UPS", "Server",
    "POS Beetle", "POS Custom",
    "POS ELO All in One", "POS NCR", "Other"
]

DEPLOYMENT_STATES = ["Functional", "Malfunctioned", "Retired"]
INCIDENT_STATES = ["Operational", "Incident"]

PROJECTS_MAP = {
    "106 JYSK": "106",
    "107 Tendam": "107",
    "108 Deichmann": "108",
    "109 Takko": "109",
    "112 Mercator-S": "112",
    "115 H&M": "115",
    "118 Metro Cash and Carry": "118",
    "119 IKEA": "119",
    "120 Peek & Cloppenburg": "120",
    "122 LPP": "122",
    "123 Decathlon": "123",
    "146 LC Waikiki": "146",
    "148 Lukoil": "148",
    "159 Emmezeta": "159",
    "179 Loreal": "179",
    "192 OMV": "192",
    "193 Lidl": "193",
}

PROJECTS_REVERSE = {v: k for k, v in PROJECTS_MAP.items()}
PROJECTS_LABELS = sorted(list(PROJECTS_MAP.keys()))

@st.cache_data
def load_existing_data():
    try:
        return pd.read_excel("data.xlsx", dtype=str).fillna("")
    except:
        return pd.DataFrame()

existing_data = load_existing_data()

def get_options(column_name, fallback=None):
    fallback = fallback or []

    if existing_data.empty or column_name not in existing_data.columns:
        return fallback

    values = (
        existing_data[column_name]
        .astype(str)
        .str.strip()
        .replace("", pd.NA)
        .dropna()
        .unique()
        .tolist()
    )

    clean_values = []

    for v in values:
        if column_name == "Project" and v in PROJECTS_REVERSE:
            clean_values.append(PROJECTS_REVERSE[v])
        else:
            clean_values.append(v)

    return sorted(list(dict.fromkeys(clean_values + fallback)))

NAME_OPTIONS = get_options("Name")
VENDOR_OPTIONS = get_options("Vendor")
MODEL_OPTIONS = get_options("Model")
TYPE_OPTIONS = get_options("Type", DEFAULT_TYPE_OPTIONS)
PROJECT_OPTIONS = get_options("Project", PROJECTS_LABELS)

def predictive_dropdown(label, options, key):
    try:
        return st.selectbox(
            label,
            options=options,
            index=None,
            placeholder="",
            accept_new_options=True,
            key=key
        )
    except TypeError:
        return st.selectbox(
            label,
            options=[""] + options,
            index=0,
            key=key
        )

devices = []
valid = True

count = st.number_input("Broj uređaja", 1, 50, 1)

for i in range(int(count)):
    st.markdown("---")
    st.subheader(f"📦 Uređaj {i+1}")

    name = predictive_dropdown("Name *", NAME_OPTIONS, key=f"name{i}")
    if not name:
        valid = False

    vendor = predictive_dropdown("Vendor", VENDOR_OPTIONS, key=f"vendor{i}")
    model = predictive_dropdown("Model", MODEL_OPTIONS, key=f"model{i}")

    type_label = predictive_dropdown("Type *", TYPE_OPTIONS, key=f"type{i}")
    if not type_label:
        valid = False

    sp = barcode_text_input("SPInventoryNumber *", key=f"sp{i}")
    sp_clean = sp.strip()

    if not sp_clean or len(sp_clean) != 7 or not (sp_clean.startswith("FS") or sp_clean.startswith("SP")):
        valid = False

    inventory = barcode_text_input("InventoryNumber", key=f"inv{i}")
    serial = barcode_text_input("SerialNumber", key=f"serial{i}")

    deployment = st.selectbox(
        "Deployment State",
        DEPLOYMENT_STATES,
        index=0,
        key=f"dep{i}"
    )

    incident = st.selectbox(
        "Incident State",
        INCIDENT_STATES,
        index=0,
        key=f"inc{i}"
    )

    project_label = predictive_dropdown(
        "Project",
        PROJECT_OPTIONS,
        key=f"proj{i}"
    )

    project_value = PROJECTS_MAP.get(
        project_label,
        project_label.split(" ")[0] if project_label else ""
    )

    devices.append({
        "Name": name,
        "Vendor": vendor,
        "Model": model,
        "Type": type_label,
        "SPInventoryNumber": sp_clean,
        "InventoryNumber": inventory,
        "SerialNumber": serial,
        "Deployment State": deployment,
        "Incident State": incident,
        "Project": project_value
    })

def set_cell(ws, cell, value):
    for merged_range in ws.merged_cells.ranges:
        if cell in merged_range:
            top_left = merged_range.start_cell.coordinate
            ws[top_left] = value
            ws[top_left].alignment = Alignment(horizontal="center", vertical="center")
            return

    ws[cell] = value
    ws[cell].alignment = Alignment(horizontal="center", vertical="center")

def prepare_df():
    df = pd.DataFrame(devices)
    df["Type"] = df["Type"].astype(str).str.replace(r"[^\w\s\-\/]", "", regex=True).str.strip()
    return df

def validate_devices(df):
    errors = {}

    try:
        existing_df = pd.read_excel("data.xlsx", dtype=str).fillna("")
    except:
        existing_df = pd.DataFrame()

    for col in ["SPInventoryNumber", "InventoryNumber", "SerialNumber"]:
        if col in existing_df.columns:
            existing_values = set(existing_df[col].astype(str).str.strip())

            for idx, val in enumerate(df[col]):
                val = str(val).strip()

                if val and val in existing_values:
                    errors.setdefault(idx, []).append(f"{col} već postoji ({val})")

    for col in ["SPInventoryNumber", "InventoryNumber", "SerialNumber"]:
        dup = df[col].astype(str).str.strip().duplicated(keep=False)

        for idx in df[dup].index:
            val = str(df.loc[idx, col]).strip()

            if val:
                errors.setdefault(idx, []).append(f"Duplikat ({col}: {val})")

    return errors

def show_errors(errors):
    st.error("❌ Pronađene greške:")
    for idx, msgs in errors.items():
        st.warning(f"Uređaj {idx + 1}: " + " | ".join(set(msgs)))

def check_before_export():
    if not valid:
        st.error("❌ Popuni obavezna polja: Name, Type i SPInventoryNumber")
        st.stop()

    df = prepare_df()
    errors = validate_devices(df)

    if errors:
        show_errors(errors)
        st.stop()

    return df

st.markdown("---")
st.subheader("⬇️ Export")

if st.button("📥 Download CMDB Excel"):
    df = check_before_export()

    output = BytesIO()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="CMDB")

    st.download_button(
        "📥 Preuzmi CMDB Excel",
        data=output.getvalue(),
        file_name="cmdb_unos.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

st.markdown("---")
st.subheader("📄 Dokument")

col_doc1, col_doc2 = st.columns(2)

with col_doc1:
    if st.button("📄 Otpremnica"):
        st.session_state.doc_type = "otpremnica"

with col_doc2:
    if st.button("📄 Prijemnica"):
        st.session_state.doc_type = "prijemnica"

if st.session_state.doc_type == "otpremnica":

    st.markdown("---")
    st.subheader("📄 Podaci za otpremnicu")

    datum = st.date_input("Datum otpremnice", value=date.today())
    iz_magacina = st.text_input("Iz magacina *", key="otp_iz_magacina")
    magacin = st.text_input("Magacin *", key="otp_magacin")

    if st.button("📥 Generiši Otpremnicu"):

        if not iz_magacina.strip():
            st.error("❌ Polje 'Iz magacina' je obavezno")
            st.stop()

        if not magacin.strip():
            st.error("❌ Polje 'Magacin' je obavezno")
            st.stop()

        df = check_before_export()

        try:
            wb = load_workbook("otpremnica_template.xlsx")
            ws = wb.active
        except:
            st.error("❌ Nije pronađen fajl: otpremnica_template.xlsx")
            st.stop()

        set_cell(ws, "G5", datum.strftime("%d.%m.%Y"))
        set_cell(ws, "B8", iz_magacina)
        set_cell(ws, "E8", magacin)

        start_row = 14

        for i, d in enumerate(devices):
            r = start_row + i

            set_cell(ws, f"B{r}", i + 1)
            set_cell(ws, f"C{r}", d["Name"])
            set_cell(ws, f"D{r}", d["Model"])
            set_cell(ws, f"E{r}", d["InventoryNumber"])
            set_cell(ws, f"F{r}", d["SerialNumber"])
            set_cell(ws, f"G{r}", d["SPInventoryNumber"])

        output = BytesIO()
        wb.save(output)

        st.download_button(
            "📥 Preuzmi Otpremnicu",
            data=output.getvalue(),
            file_name="otpremnica.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

if st.session_state.doc_type == "prijemnica":

    st.markdown("---")
    st.subheader("📄 Podaci za prijemnicu")

    datum = st.date_input("Datum prijemnice", value=date.today())
    iz_magacina = st.text_input("Iz magacina *", key="pri_iz_magacina")
    magacin = st.text_input("Magacin *", key="pri_magacin")

    if st.button("📥 Generiši Prijemnicu"):

        if not iz_magacina.strip():
            st.error("❌ Polje 'Iz magacina' je obavezno")
            st.stop()

        if not magacin.strip():
            st.error("❌ Polje 'Magacin' je obavezno")
            st.stop()

        df = check_before_export()

        try:
            wb = load_workbook("prijemnica_template.xlsx")
            ws = wb.active
        except:
            st.error("❌ Nije pronađen fajl: prijemnica_template.xlsx")
            st.stop()

        set_cell(ws, "G5", datum.strftime("%d.%m.%Y"))
        set_cell(ws, "B8", iz_magacina)
        set_cell(ws, "E8", magacin)

        start_row = 14

        for i, d in enumerate(devices):
            r = start_row + i

            set_cell(ws, f"B{r}", i + 1)
            set_cell(ws, f"C{r}", d["Name"])
            set_cell(ws, f"D{r}", d["Model"])
            set_cell(ws, f"E{r}", d["InventoryNumber"])
            set_cell(ws, f"F{r}", d["SerialNumber"])
            set_cell(ws, f"G{r}", d["SPInventoryNumber"])

        output = BytesIO()
        wb.save(output)

        st.download_button(
            "📥 Preuzmi Prijemnicu",
            data=output.getvalue(),
            file_name="prijemnica.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
